from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# ���� ���� ���� API
@app.route('/create_excel', methods=['POST'])
def create_excel_file():
    file_name = request.json.get('file_name', 'user_data.xlsx')

    if os.path.exists(file_name):
        return jsonify({"message": f"{file_name} ������ �̹� �����մϴ�."}), 400

    # ���ο� ���� ��ũ�� ����
    wb = Workbook()
    ws = wb.active

    # ��� �߰�
    ws.append(["�г���", "���̵�", "��й�ȣ", "�ְ� ���"])

    # ���� ����
    wb.save(file_name)
    return jsonify({"message": f"{file_name} ������ �����߽��ϴ�."}), 201

# ����� ���� �߰� API
@app.route('/add_user', methods=['POST'])
def add_user_to_excel():
    data = request.json
    file_name = data.get('file_name', 'user_data.xlsx')
    nickname = data.get('nickname')
    user_id = data.get('user_id')
    password = data.get('password')
    high_score = data.get('high_score')

    if not os.path.exists(file_name):
        return jsonify({"message": f"{file_name} ������ �����ϴ�. ���� ������ �����ϼ���."}), 404

    # ���� ���� ����
    wb = load_workbook(file_name)
    ws = wb.active

    # ������ �߰�
    ws.append([nickname, user_id, password, high_score])

    # ���� ����
    wb.save(file_name)
    return jsonify({
        "message": "����� ������ �߰��Ǿ����ϴ�.",
        "user": {
            "nickname": nickname,
            "user_id": user_id,
            "high_score": high_score
        }
    }), 200

# ���� �Լ�
if __name__ == '__main__':
    app.run(debug=True)
